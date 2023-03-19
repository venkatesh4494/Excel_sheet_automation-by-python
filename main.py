import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.chart import BarChart,Reference
from openpyxl.styles import Font

df=pd.read_excel("/Users/venky/PycharmProjects/excel_auto/supermarket_sales.xlsx")
df=df[["Gender","Product line","Total"]]
table = pd.pivot_table(df, values='Total',index="Gender",
                    columns=["Product line"], aggfunc=np.sum)
table.to_excel("pivot_table.xlsx","Report")
wb=load_workbook("/Users/venky/PycharmProjects/excel_auto/pivot_table.xlsx")
sheet=wb["Report"]
#sheet['A1']="Sales Report"
#sheet['A2']="January"
#sheet['A1'].font=Font("Arial",bold=True,size=20)
#sheet['A2'].font=Font("Arial",bold=True,size=10)
min_row=wb.active.min_row
max_row=wb.active.max_row
min_col=wb.active.min_column
max_col=wb.active.max_column
barchart=BarChart()
data=Reference(sheet,
               min_col=min_col+1,
               max_col=max_col,
                 min_row=min_row,
                 max_row=max_row
)
categories = Reference(sheet,
                       min_col=min_col,
                       max_col=min_col,
                       min_row=min_row+1,
                       max_row=max_row)

barchart.add_data(data,titles_from_data=True)
barchart.set_categories(categories)

sheet.add_chart(barchart,"B7")
barchart.title="Sales by Product line"
barchart.style=5
wb.save("barchart.xlsx")
a=0