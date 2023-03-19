import pandas as pd
import numpy as np
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.chart import BarChart,Reference
from openpyxl.utils import get_column_letter
import os
import sys

app_path=os.path.dirname(sys.executable)

month=input('Enter the month:')

input_path=os.path.join(app_path,"pivot_table.xlsx")
wb=load_workbook(input_path)
sheet=wb["Report"]

#sheet['A1']="Sales Report"
#sheet['A2']="January"
#sheet['A1'].font=Font("Arial",bold=True,size=20)
#sheet['A2'].font=Font("Arial",bold=True,size=10)

min_row=wb.active.min_row
max_row=wb.active.max_row
min_col=wb.active.min_column
max_col=wb.active.max_column

barchart=BarChart()

data=Reference(sheet,
               min_col=min_col+1,
               max_col=max_col,
                 min_row=min_row,
                 max_row=max_row
)
categories = Reference(sheet,
                       min_col=min_col,
                       max_col=min_col,
                       min_row=min_row+1,
                       max_row=max_row)

barchart.add_data(data,titles_from_data=True)
barchart.set_categories(categories)

sheet.add_chart(barchart,"B7")
barchart.title="Sales by Product line"
barchart.style=5

for i in range(min_col+1,max_col+1):
    letter=get_column_letter(i)
    sheet[f'{letter}{max_row+1}']=f'=SUM({letter}{min_row + 1}:{letter}{max_row})'
    sheet[f'{letter}{max_row + 1}'].style = 'Currency'

filename=f'report_{month}.xlsx'
output_path=os.path.join(app_path,filename)
wb.save(output_path)